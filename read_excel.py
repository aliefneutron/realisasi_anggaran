import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook(r'C:\Users\ALIEF COMPUTER\Downloads\Realisasi Keuangan\Realisasi Keuangan\data\REALISASI BELANJA DINKES.xlsx', read_only=True, data_only=True)
ws = wb.active

print(f'Sheet name: {ws.title}')
print('\nColumns (first row):')
first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
print(first_row)

print('\nFirst 15 data rows:')
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=16, values_only=True)):
    print(f'Row {i+2}: {row}')

wb.close()
